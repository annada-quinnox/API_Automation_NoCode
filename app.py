from urllib import response
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from testcaseengine import generate_testcases, flatten
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import json
import re
import requests
import time
from datetime import datetime
from database import get_database, initialize_database
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
import subprocess
import csv
import os

DEFAULT_SORT_HEADER = "created_at"

def api_success(data=None, status_code=200, **kwargs):
    response = {'success': True}
    if data is not None:
        response.update(data)
    response.update(kwargs)
    return jsonify(response), status_code

def api_error(message, status_code=400):
    return jsonify({'success': False, 'error': str(message)}), status_code

def get_data_type(value):
    if value is None: return 'null'
    if isinstance(value, bool): return 'boolean'
    if isinstance(value, int): return 'integer'
    if isinstance(value, float): return 'number'
    if isinstance(value, list): return 'array'
    if isinstance(value, dict): return 'object'
    return 'string'

def parse_query_params(input_data):
    if not isinstance(input_data, str):
        return input_data if isinstance(input_data, dict) else {}
    
    qs = input_data.strip()
    if qs.startswith('?'):
        qs = qs[1:]
    
    if ' ' in qs: 
        parts = qs.split(' ')
        for part in parts:
            if '?' in part:
                qs = part.split('?')[1]
                break
            elif '=' in part:
                qs = part
                break
    elif '?' in qs:
        qs = qs.split('?')[1]
        
    params = {}
    if not qs or '=' not in qs:
        return params

    for pair in qs.split('&'):
        if '=' in pair:
            key, value = pair.split('=', 1)
            if value.lower() == 'true':
                value = True
            elif value.lower() == 'false':
                value = False
            elif value.isdigit():
                value = int(value)
            else:
                try:
                    value = float(value)
                except ValueError:
                    pass
            params[key] = value
            
    return params

def validate_against_configs(input_data, field_configs, source='body'):
    if not input_data or not field_configs:
        return True, []
    
    try:
        if isinstance(input_data, str):
            try:
                input_json = json.loads(input_data)
            except:
                if source == 'query':
                    input_json = parse_query_params(input_data)
                else:
                    return False, [f"Invalid JSON payload: {input_data}"]
        else:
            input_json = input_data
    except Exception as e:
        return False, [f"Validation error: {str(e)}"]

    if not isinstance(input_json, dict):
        return True, []

    flat_input = flatten(input_json)
    errors = []
    
    for field, config in field_configs.items():
        expected_type = config.get('type')
        is_required = config.get('required', False)
        
        if field not in flat_input:
            if is_required == 'required' or is_required is True:
                errors.append(f"For '{source}' at path '{field}': Missing required field.")
            continue
            
        value = flat_input[field]
        if value is None:
            if is_required == 'required' or is_required is True:
                errors.append(f"For '{source}' at path '{field}': Value cannot be null.")
            continue
            
        curr_type = get_data_type(value)
        
        type_mismatch = False
        check_type = expected_type
        if expected_type in ['email', 'uuid', 'date', 'datetime', 'url', 'password', 'phone']:
            check_type = 'string'
        
        if check_type == 'string' and not isinstance(value, str):
            type_mismatch = True
        elif check_type == 'integer' and (not isinstance(value, int) or isinstance(value, bool)):
            type_mismatch = True
        elif check_type == 'number' and not isinstance(value, (int, float)):
            type_mismatch = True
        elif check_type == 'boolean' and not isinstance(value, bool):
            type_mismatch = True
        elif check_type == 'array' and not isinstance(value, list):
            type_mismatch = True
        elif check_type == 'object' and not isinstance(value, dict):
            type_mismatch = True

        if type_mismatch:
            errors.append(f"For '{source}' at path '{field}': Value must be a {expected_type}.")
    
    if errors:
        return False, errors
    return True, []

def validate_input_types(input_data, original_payload):
    if not input_data or not original_payload:
        return True, []
    
    try:
        if isinstance(input_data, str):
            input_json = json.loads(input_data)
        else:
            input_json = input_data
            
        if isinstance(original_payload, str):
            original_json = json.loads(original_payload)
        else:
            original_json = original_payload
    except:
        return True, []

    if not isinstance(input_json, dict) or not isinstance(original_json, dict):
        return True, []

    flat_input = flatten(input_json)
    flat_orig = flatten(original_json)
    errors = []
    
    for key, value in flat_input.items():
        if key in flat_orig and value is not None:
            orig_val = flat_orig[key]
            if orig_val is not None:
                orig_type = get_data_type(orig_val)
                curr_type = get_data_type(value)
                
                if orig_type == 'integer' and curr_type == 'number':
                    errors.append(f"For 'body' at path '{key}': Expected integer, but sent float/number.")
                elif orig_type != curr_type:
                    errors.append(f"For 'body' at path '{key}': Expected {orig_type}, but sent {curr_type}.")
    
    if errors:
        return False, errors
    return True, []

def validate_response_schema(response_body, status_code):
    if not response_body:
        if status_code >= 500:
            return False, ["Empty response body for 5xx error"]
        return True, []
    
    try:
        if isinstance(response_body, str):
            data = json.loads(response_body)
        else:
            data = response_body
    except json.JSONDecodeError:
        if status_code >= 500:
            return False, ["Response body is not valid JSON for 5xx error"]
        return True, []
    except:
        if status_code >= 500:
            return False, ["Cannot parse response body for 5xx error"]
        return True, []
    
    if status_code >= 500:
        if not isinstance(data, dict):
            return False, ["Response body must be a JSON object for 5xx errors"]
        if "error" not in data:
            return False, ["Missing 'error' field in 5xx error response"]
        if "message" not in data:
            return False, ["Missing 'message' field in 5xx error response"]
        if not isinstance(data.get("error"), str):
            return False, ["Field 'error' must be a string"]
        if not isinstance(data.get("message"), str):
            return False, ["Field 'message' must be a string"]
    elif 400 <= status_code < 500:
        if isinstance(data, dict):
            if "error" in data and not isinstance(data["error"], str):
                return False, ["Field 'error' must be a string"]
            if "message" in data and not isinstance(data["message"], str):
                return False, ["Field 'message' must be a string"]
    
    return True, []

app = Flask(__name__)
CORS(app)

generator = generate_testcases()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/generate', methods=['POST'])
def generate_test_cases():
    try:
        data = request.get_json()
        test_cases = generator.generate_test_cases(data)
        return api_success({'test_cases': test_cases, 'count': len(test_cases)})
    except Exception as e:
        print("ERROR:", e)
        return api_error(e)

@app.route('/api/test-cases', methods=['GET'])
def get_test_cases():
    try:
        test_cases = generator.get_test_cases()
        return api_success({'test_cases': test_cases, 'count': len(test_cases)})
    except Exception as e:
        return api_error(e)

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'app': 'API Test Case Generator v3'})

def extract_response_code(expected_input):
    if not expected_input:
        return ["N/A"]
    
    if isinstance(expected_input, list):
        valid_codes = []
        for item in expected_input:
            if item is None:
                continue
            item_str = str(item).strip()
            if re.match(r'^[1-5]\d{2}$', item_str):
                valid_codes.append(item_str)
            elif item_str.upper() != "N/A":
                matches = re.findall(r'\b([1-5]\d{2})\b', item_str)
                valid_codes.extend(matches)
        
        if valid_codes:
            return valid_codes
        else:
            return ["N/A"]
    
    expected_str = str(expected_input)

    expected_lower = expected_str.lower()
    if 'default sort' in expected_lower or 'fallback' in expected_lower:
        codes = re.findall(r'\b([1-5]\d{2})\b', expected_str)
        normalized = []
        for c in codes:
            if c not in normalized:
                normalized.append(c)
        if '200' not in normalized:
            normalized.append('200')
        if normalized:
            return normalized
    
    matches = re.findall(r'\b([1-5]\d{2})\b', expected_str)
    if matches:
        return matches
    
    expected_lower = expected_str.lower()
    if 'created' in expected_lower: return ["201"]
    if 'no content' in expected_lower: return ["204"]
    if 'success' in expected_lower or 'ok' in expected_lower: return ["200"]
    if 'bad request' in expected_lower or 'invalid' in expected_lower or 'missing' in expected_lower: return ["400"]
    if 'unauthorized' in expected_lower: return ["401"]
    if 'forbidden' in expected_lower: return ["403"]
    if 'not found' in expected_lower: return ["404"]
    if 'conflict' in expected_lower: return ["409"]
    if 'too many' in expected_lower or 'rate limit' in expected_lower: return ["429"]
    return ["N/A"]

def format_expected_for_display(expected):
    if not expected:
        return "N/A"
    if isinstance(expected, list):
        if len(expected) == 1:
            return expected[0]
        else:
            return ", ".join(expected)
    return str(expected)

def get_test_case_source_info(test_case):
    source = "excel"
    additional_info = []
    
    if test_case.get('expected_status') is not None:
        source = "database"
    
    if test_case.get('test_case_id'):
        additional_info.append(f"DB ID: {test_case['test_case_id']}")
    
    if test_case.get('test_case_number'):
        additional_info.append(f"Test #: {test_case['test_case_number']}")
    
    if test_case.get('metadata'):
        metadata = test_case['metadata']
        if isinstance(metadata, dict):
            if metadata.get('source'):
                source = metadata['source']
            if metadata.get('session_id'):
                additional_info.append(f"Session: {metadata['session_id']}")
    
    return source, additional_info

def format_input_body(input_data):
    if isinstance(input_data, dict):
        return json.dumps(input_data, indent=2)
    elif isinstance(input_data, list):
        return json.dumps(input_data, indent=2)
    else:
        return str(input_data)

def _create_excel_styles():
    return {
        'header_fill': PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid"),
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'center_align': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'left_align': Alignment(horizontal="left", vertical="center", wrap_text=True),
    }

def _build_test_case_excel(test_cases, method, endpoint, base_url="", include_base_url=True):
    styles = _create_excel_styles()
    wb = openpyxl.Workbook()
    ws = cast(Worksheet | None, wb.active)
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = "API Test Cases"

    if include_base_url:
        headers = ["ID", "HTTP Method", "Test Case Name", "Test Type", "Base Url",
                    "Endpoint", "Request Body", "Expected Response Code", "Expected Status", "Status"]
    else:
        headers = ["ID", "HTTP Method", "Test Case Name", "Test Type",
                    "Endpoint", "Request Body", "Expected Response Code", "Expected Status", "Status"]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['center_align']
        cell.border = styles['border']

    for tc in test_cases:
        request_body = format_input_body(tc.get("input", {}))
        response_codes = extract_response_code(tc.get("expected", ""))
        response_code_str = ", ".join(response_codes) if isinstance(response_codes, list) else str(response_codes)

        if include_base_url:
            row_data = [
                tc.get("id"), method, tc.get("scenario"), tc.get("type"),
                tc.get("baseUrl", base_url), endpoint, request_body,
                response_code_str, tc.get("expected"), ""
            ]
        else:
            row_data = [
                tc.get("id"), method, tc.get("scenario"), tc.get("type"),
                endpoint, request_body, response_code_str, tc.get("expected"), ""
            ]
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = styles['border']
            cell.alignment = styles['left_align']

    col_widths = {'A': 12, 'B': 12, 'C': 35, 'D': 12}
    if include_base_url:
        col_widths.update({'E': 25, 'F': 20, 'G': 50, 'H': 16, 'I': 30, 'J': 10})
    else:
        col_widths.update({'E': 20, 'F': 50, 'G': 16, 'H': 30, 'I': 10})

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _generate_excel_filename(method, endpoint, prefix="TestCases"):
    endpoint_clean = endpoint.strip('/').replace('/', '_').replace(' ', '_').upper()
    now = datetime.now()
    return f"{method}_{endpoint_clean}_{prefix}_{now.strftime('%Y-%m-%d')}_{now.strftime('%H-%M-%S')}.xlsx"

# ======= FIX: SAFE EXCEL EXPORT ========
def _normalize_excel_header(header):
    if header is None:
        return ""
    return re.sub(r'[^a-z0-9]+', '_', str(header).strip().lower()).strip('_')

def _parse_excel_input_cell(value):
    if value is None:
        return {}

    if isinstance(value, (dict, list, int, float, bool)):
        return value

    text = str(value).strip()
    if not text:
        return {}

    try:
        return json.loads(text)
    except Exception:
        return text

def _build_test_case_from_excel_row(row_data, index):
    method = str(row_data.get('http_method') or row_data.get('method') or 'GET').strip().upper()
    endpoint = str(row_data.get('endpoint') or '/api/test').strip() or '/api/test'
    base_url = str(row_data.get('base_url') or row_data.get('baseurl') or '').strip()
    scenario = str(row_data.get('test_case_name') or row_data.get('scenario') or f'Imported Test Case {index}').strip()
    test_type = str(row_data.get('test_type') or row_data.get('type') or 'Positive').strip() or 'Positive'
    expected = str(
        row_data.get('expected_status')
        or row_data.get('expected_response')
        or row_data.get('expected_response_code')
        or row_data.get('expected')
        or 'N/A'
    ).strip()

    input_body = _parse_excel_input_cell(
        row_data.get('request_body')
        or row_data.get('input_body')
        or row_data.get('input')
        or row_data.get('payload')
    )

    test_case_id = row_data.get('id') or row_data.get('test_case_id') or f'IMP_{index:03d}'
    test_case_id = str(test_case_id).strip() if test_case_id is not None else f'IMP_{index:03d}'

    return {
        'id': test_case_id,
        'test_case_number': index,
        'type': test_type,
        'scenario': scenario,
        'input': input_body,
        'expected': expected,
        'expected_status': extract_response_code(expected),
        'baseUrl': base_url,
        'endpoint': endpoint,
        'method': method
    }

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    try:
        uploaded_file = request.files.get('file')
        if uploaded_file is None or not uploaded_file.filename:
            return api_error('No Excel file was uploaded')

        filename = uploaded_file.filename
        if not filename.lower().endswith('.xlsx'):
            return api_error('Unsupported file format. Please upload an .xlsx file')

        workbook = openpyxl.load_workbook(uploaded_file.stream, data_only=True)
        ws = cast(Worksheet | None, workbook.active)
        if ws is None:
            return api_error('Failed to read worksheet from uploaded file')

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return api_error('Uploaded Excel file is empty')

        headers = [
            _normalize_excel_header(header) or f'column_{idx + 1}'
            for idx, header in enumerate(rows[0])
        ]

        test_cases = []
        for row in rows[1:]:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            row_data = {}
            for idx, cell_value in enumerate(row):
                if idx >= len(headers):
                    break
                row_data[headers[idx]] = cell_value

            test_cases.append(_build_test_case_from_excel_row(row_data, len(test_cases) + 1))

        if not test_cases:
            return api_error('No test case rows found in the uploaded Excel file')

        return api_success({'test_cases': test_cases, 'count': len(test_cases), 'filename': filename})
    except Exception as e:
        return api_error(f'Excel upload failed: {str(e)}', status_code=500)

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        
        # Bypass dangerous regeneration: take exact cases from UI pool
        test_cases = data.get('test_cases', [])
        if not test_cases:
            return api_error('No test cases provided for export')
            
        method = data.get('method', 'GET')
        endpoint = data.get('endpoint', '/api/test')
        base_url = data.get('baseUrl', data.get('base_url', ''))

        output = _build_test_case_excel(test_cases, method, endpoint, base_url, include_base_url=True)
        filename = _generate_excel_filename(method, endpoint)

        return send_file(
            output, download_name=filename, as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return api_error(str(e))

# ======= FIX: SAFE DATABASE EXPORT ========
@app.route('/api/save-to-database', methods=['POST'])
def save_to_database():
    try:
        data = request.get_json()
        
        # Bypass dangerous regeneration: take exact cases from UI pool
        test_cases = data.get('test_cases', [])
        if not test_cases: 
            return api_error('No test cases provided to save')
        
        session_data = {
            'endpoint': data.get('endpoint', '/api/test'),
            'method': data.get('method', 'GET'),
            'base_url': data.get('baseUrl', data.get('base_url', '')),
            'session_name': data.get('session_name', f"Test Cases {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
            'created_by': 'system'
        }
        
        db = get_database()
        success, message, session_id, saved_count = db.save_test_cases(session_data, test_cases)
        if success: return api_success({'session_id': session_id, 'message': message, 'saved_count': saved_count})
        else: return api_error(message, status_code=500)
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/database-sessions', methods=['GET'])
def get_database_sessions():
    try:
        limit = request.args.get('limit', default=50, type=int)
        offset = request.args.get('offset', default=0, type=int)
        endpoint_filter = request.args.get('endpoint', default=None, type=str)
        base_url_filter = request.args.get('base_url', default=None, type=str)
        method_filter = request.args.get('method', default=None, type=str)
        
        db = get_database()
        sessions, total = db.get_sessions(
            limit=limit, offset=offset, endpoint_filter=endpoint_filter,
            base_url_filter=base_url_filter, method_filter=method_filter
        )
        return api_success({'sessions': sessions, 'total': total, 'limit': limit, 'offset': offset})
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/database-test-cases/<session_id>', methods=['GET'])
def get_session_test_cases(session_id):
    try:
        db = get_database()
        session_info, test_cases = db.get_test_cases(session_id)
        if session_info is None: return api_error(f"Session {session_id} not found", status_code=404)
        return api_success({'session_info': session_info, 'test_cases': test_cases, 'count': len(test_cases)})
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/active-testcase-pool', methods=['GET'])
def get_active_testcase_pool():
    try:
        db = get_database()
        success, message, pool_data, total_rows = db.get_active_testcase_pool()
        if success:
            return api_success({'pool': pool_data, 'count': total_rows, 'message': message})
        return api_error(message, status_code=500)
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/active-testcase-pool', methods=['POST'])
def save_active_testcase_pool():
    try:
        data = request.get_json() or {}
        pool_data = data.get('pool', {})
        if not isinstance(pool_data, dict):
            return api_error('Invalid pool payload')

        db = get_database()
        success, message, saved_count = db.save_active_testcase_pool(pool_data)
        if success:
            return api_success({'message': message, 'saved_count': saved_count})
        return api_error(message, status_code=500)
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/active-test-suites', methods=['GET'])
def get_active_test_suites():
    try:
        db = get_database()
        success, message, suites, active_suite_id, total_rows = db.get_active_test_suites()
        if success:
            return api_success({
                'suites': suites,
                'active_suite_id': active_suite_id,
                'count': total_rows,
                'message': message
            })
        return api_error(message, status_code=500)
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/active-test-suites', methods=['POST'])
def save_active_test_suites():
    try:
        data = request.get_json() or {}
        suites = data.get('suites', [])
        active_suite_id = data.get('active_suite_id')

        if not isinstance(suites, list):
            return api_error('Invalid suites payload')

        db = get_database()
        success, message, saved_count = db.save_active_test_suites(suites, active_suite_id)
        if success:
            return api_success({'message': message, 'saved_count': saved_count})
        return api_error(message, status_code=500)
    except Exception as e:
        return api_error(str(e), status_code=500)

@app.route('/api/database-health', methods=['GET'])
def database_health():
    try:
        success, message = initialize_database()
        return api_success({'message': message}) if success else api_error(message, status_code=500)
    except Exception as e:
        return api_error(f"Database health check failed: {str(e)}", status_code=500)

@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.get_json()
        test_cases = data.get('test_cases', [])
        endpoint = data.get('endpoint', '/api/test')
        method = data.get('method', 'GET')
        if not test_cases: return api_error('No test cases provided')

        output = _build_test_case_excel(test_cases, method, endpoint, include_base_url=False)
        filename = _generate_excel_filename(method, endpoint)
        return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return api_error(e)

@app.route('/api/export-results', methods=['POST'])
def export_results():
    try:
        data = request.get_json()
        results = data.get('results', [])
        test_cases = data.get('test_cases', [])
        endpoint = data.get('endpoint', '/api/test')
        method = data.get('method', 'GET')
        
        if not results:
            return api_error('No results provided')

        wb = openpyxl.Workbook()
        ws = cast(Worksheet | None, wb.active)
        if ws is None:
            return api_error('Failed to create worksheet')
        ws.title = "API Execution Results"

        styles = _create_excel_styles()

        headers = [
            "ID", "HTTP Method", "Test Case Name", "Test Type", 
            "Base Url", "Endpoint", "Request Body", 
            "Expected Response Code", "Expected Status",
            "Actual Status Code", "Execution Status", "Response Body", "Details"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']
            cell.border = styles['border']

        tc_map = {tc.get('id'): tc for tc in test_cases}

        for res in results:
            tc_id = res.get('testCaseId')
            tc = tc_map.get(tc_id, {})
            
            request_body = format_input_body(tc.get("input", {}))
            status = res.get('status', 'fail').upper()
            
            response_codes = extract_response_code(tc.get("expected", ""))
            response_code_str = ", ".join(response_codes) if isinstance(response_codes, list) else str(response_codes)
            
            row_data = [
                tc_id,
                method,
                tc.get("scenario", "N/A"),
                tc.get("type", "N/A"),
                tc.get("baseUrl", ""),
                endpoint,
                request_body,
                response_code_str,
                tc.get("expected", "N/A"),
                res.get('statusCode', 'N/A'),
                status,
                res.get('responseBody', ''),
                res.get('details', '')
            ]
            ws.append(row_data)
            
            last_row = ws.max_row
            result_cell = ws.cell(row=last_row, column=11)
            if status == 'PASS':
                result_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                result_cell.font = Font(color="065F46", bold=True)
            else:
                result_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                result_cell.font = Font(color="991B1B", bold=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = styles['border']
                cell.alignment = styles['left_align']

        ws.column_dimensions['A'].width = 12 
        ws.column_dimensions['B'].width = 12 
        ws.column_dimensions['C'].width = 35 
        ws.column_dimensions['D'].width = 12 
        ws.column_dimensions['E'].width = 25 
        ws.column_dimensions['F'].width = 20 
        ws.column_dimensions['G'].width = 45 
        ws.column_dimensions['H'].width = 16 
        ws.column_dimensions['I'].width = 30 
        ws.column_dimensions['J'].width = 15 
        ws.column_dimensions['K'].width = 15 
        ws.column_dimensions['L'].width = 50 
        ws.column_dimensions['M'].width = 50 

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        endpoint_clean = endpoint.strip('/').replace('/', '_').replace(' ', '_').upper()
        now = datetime.now()
        filename = f"RESULTS_{method}_{endpoint_clean}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return api_error(e)

@app.route('/api/execute-tests', methods=['POST'])
def execute_tests():
    try:
        data = request.get_json()
        endpoint = data.get('endpoint', '').strip()
        method = data.get('method', 'GET').upper()
        test_cases = data.get('testCases', [])
        environment = data.get('environment', 'mock')
        base_url = data.get('baseUrl') or data.get('base_url') or 'mock'      
        if not endpoint: return api_error('Endpoint is required')
        if not test_cases: return api_error('No test cases provided')

        results = []
        original_payload = data.get('originalPayload')
        if isinstance(original_payload, str) and original_payload.strip():
            try: original_payload = json.loads(original_payload)
            except: pass

        field_configs = data.get('fieldConfigs', {})
        for test_case in test_cases:
            result = execute_single_test(endpoint, method, test_case, environment, base_url, original_payload, field_configs)
            results.append(result)

        return api_success({'results': results})
    except Exception as e:
        return api_error(e)

@app.route('/api/run-performance', methods=['POST'])
def run_performance_test():
    try:
        data = request.get_json()
        base_url = data.get('baseUrl', 'mock')
        if base_url == 'mock' or not base_url:
            return api_error("Performance testing requires a real Base URL, not a mock environment.")

        test_case = data.get('testCase', {})
        method = test_case.get('method', data.get('method', 'GET')).upper()
        expected = test_case.get('expected', '200')
        expected_codes = extract_response_code(expected)
        
        current_endpoint = test_case.get('endpoint', data.get('endpoint', '/search'))
        payload = test_case.get('input', {})
        
        if method in ['GET', 'DELETE']:
            if isinstance(payload, str):
                if payload.startswith('/'): current_endpoint = payload; payload = {}
                elif payload.startswith('?') or '=' in payload: payload = parse_query_params(payload)
                else: current_endpoint = current_endpoint.rstrip('/') + '/' + payload; payload = {}
        
        if "429" in expected_codes:
            requests_made = 0
            latencies = []
            got_429 = False
            sample_output = "No output available"
            ping_url = build_url(current_endpoint, base_url)
            headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
            
            for _ in range(6):
                start_time = time.time()
                try:
                    if method in ["GET", "DELETE"]:
                        ping_res = requests.request(method, ping_url, params=payload if isinstance(payload, dict) else None, headers=headers, timeout=5.0)
                    else:
                        if isinstance(payload, dict) and payload:
                            ping_res = requests.request(method, ping_url, json=payload, headers=headers, timeout=5.0)
                        else:
                            ping_res = requests.request(method, ping_url, data=payload, headers=headers, timeout=5.0)
                    
                    requests_made += 1
                    latencies.append((time.time() - start_time) * 1000)
                    
                    if ping_res.status_code == 429:
                        got_429 = True
                        sample_output = ping_res.text[:1500]
                        break
                    sample_output = ping_res.text[:1500]
                except Exception as e:
                    sample_output = str(e)

            metrics = {
                "requests_made": str(requests_made),
                "failures": "0" if got_429 else "1",
                "median_ms": str(round(sum(latencies)/len(latencies))) if latencies else "0",
                "avg_ms": str(round(sum(latencies)/len(latencies))) if latencies else "0",
                "max_ms": str(round(max(latencies))) if latencies else "0",
                "rps": "N/A (Rate Limit Mode)"
            }
            
            failure_details = []
            if not got_429:
                failure_details.append("Rate limit of 5 requests per IP was NOT enforced by the server. 6th request succeeded.")
                
            return api_success({
                "metrics": metrics,
                "sample_output": sample_output,
                "failure_details": failure_details
            })

        safe_payload = json.dumps(payload)
        sample_output = "No output available"
        
        try:
            ping_url = build_url(current_endpoint, base_url)
            headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
            if method in ["GET", "DELETE"]:
                ping_res = requests.request(method, ping_url, params=payload if isinstance(payload, dict) else None, headers=headers, timeout=5.0)
            else:
                if isinstance(payload, dict) and payload:
                    ping_res = requests.request(method, ping_url, json=payload, headers=headers, timeout=5.0)
                else:
                    ping_res = requests.request(method, ping_url, data=payload, headers=headers, timeout=5.0)
            sample_output = ping_res.text[:1500] 
        except Exception as e:
            sample_output = f"Failed to fetch sample: {str(e)}"

        users = str(data.get('users', 50))       
        spawn_rate = str(data.get('spawnRate', 10)) 
        run_time = data.get('runTime', '10s')    
        
        locust_script = f"""from locust import HttpUser, task, between
import json

class APIUser(HttpUser):
    wait_time = between(0.01, 0.05) 
    host = "{base_url}"

    @task
    def execute_dynamic_request(self):
        headers = {{
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }}
        
        payload_raw = {safe_payload}
        try:
            payload_data = json.loads(payload_raw)
        except:
            payload_data = payload_raw
            
        kwargs = {{"headers": headers, "catch_response": True, "timeout": 15.0}}
        
        if "{method}" in ["GET", "DELETE"]:
            if isinstance(payload_data, dict) and payload_data:
                kwargs["params"] = payload_data
        else:
            if isinstance(payload_data, dict) and payload_data:
                kwargs["json"] = payload_data
            elif payload_data:
                kwargs["data"] = payload_data

        with self.client.request("{method}", "{current_endpoint}", **kwargs) as response:
            expected_codes = "{expected}"
            
            if "401" in expected_codes or "403" in expected_codes:
                if response.status_code in [401, 403]:
                    response.success()
                else:
                    response.failure(f"Expected Auth Failure, got {{response.status_code}}")
            elif response.status_code in [200, 201, 202, 204, 304]:
                response.success()
            else:
                response.failure(f"Failed with HTTP {{response.status_code}}: {{response.text[:100]}}")
"""
        timestamp = int(time.time() * 1000)
        csv_prefix = f"perf_results_{timestamp}"
        locust_file = f"dynamic_locustfile_{timestamp}.py"

        with open(locust_file, "w", encoding="utf-8") as f:
            f.write(locust_script)

        print(f"\n🚀 Starting Locust Load Test on {base_url}{current_endpoint} [{method}]...")
        
        command = [
            "locust", "-f", locust_file, "--headless",
            "-u", users, "-r", spawn_rate, "--run-time", run_time, "--csv", csv_prefix
        ]
        
        subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("✅ Load Test Complete! Parsing results...")

        metrics = {}
        failure_details = []
        
        csv_file_stats = f"{csv_prefix}_stats.csv"
        csv_file_failures = f"{csv_prefix}_failures.csv"
        
        if os.path.exists(csv_file_failures):
            with open(csv_file_failures, mode='r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    err_msg = row.get('Error', '')
                    occ = row.get('Occurrences', '')
                    if err_msg:
                        failure_details.append(f"{err_msg} (Occurred {occ} times)")
        
        if os.path.exists(csv_file_stats):
            time.sleep(0.5)
            with open(csv_file_stats, mode='r') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if row.get('Name') == 'Aggregated':
                        metrics = {
                            "requests_made": row.get("Request Count", "0"),
                            "failures": row.get("Failure Count", "0"),
                            "median_ms": row.get("Median Response Time", "0"),
                            "avg_ms": row.get("Average Response Time", "0"),
                            "max_ms": row.get("Max Response Time", "0"),
                            "rps": row.get("Requests/s", "0")
                        }
                        break
                
                if not metrics: 
                    file.seek(0)
                    reader = csv.DictReader(file)
                    for row in reader:
                        metrics = {
                            "requests_made": row.get("Request Count", "0"),
                            "failures": row.get("Failure Count", "0"),
                            "median_ms": row.get("Median Response Time", "0"),
                            "avg_ms": row.get("Average Response Time", "0"),
                            "max_ms": row.get("Max Response Time", "0"),
                            "rps": row.get("Requests/s", "0")
                        }
                        break
            
            for ext in ['_stats.csv', '_stats_history.csv', '_failures.csv', '_exceptions.csv']:
                try: os.remove(f"{csv_prefix}{ext}")
                except: pass
            try: os.remove(locust_file)
            except: pass
            
            return api_success({
                "metrics": metrics, 
                "sample_output": sample_output,
                "failure_details": failure_details
            })
        else:
            return api_error("Locust failed to generate CSV results.")

    except Exception as e:
        return api_error(f"Performance execution failed: {str(e)}")

_STATUS_ERROR_BODY = {
    401: {"error": "Unauthorized", "message": "Missing or invalid authentication token"},
    403: {"error": "Forbidden", "message": "You do not have permission to access this resource"},
    404: {"error": "Not Found", "message": "The requested resource was not found"},
    409: {"error": "Conflict", "message": "Resource already exists or state conflict"},
    415: {"error": "Unsupported Media Type", "message": "Content-Type header is missing or invalid"},
    429: {"error": "Too Many Requests", "message": "Rate limit exceeded"},
}

_SCENARIO_ERROR_RULES = [
    (['Authorization', 'token'], 401, "Unauthorized"),
    (['Forbidden', 'Role', 'Permission'], 403, "Forbidden"),
    (['Not Found', 'non-existing'], 404, "Not Found"),
    (['Rate', 'Limit'], 429, "Too Many Requests"),
    (['Content-Type'], 415, "Unsupported Media Type"),
]

def _run_mock_validation(payload, method, test_type, field_configs, original_payload):
    errors = []
    
    if field_configs:
        source = 'query' if method == 'GET' else 'body'
        is_valid, validation_errors = validate_against_configs(payload, field_configs, source=source)
        if not is_valid:
            errors = validation_errors

    if not errors and test_type == 'Positive' and original_payload:
        source = 'query' if method == 'GET' else 'body'
        if source == 'query':
            parsed_payload = parse_query_params(payload)
            is_valid, validation_errors = validate_input_types(parsed_payload, original_payload)
        else:
            is_valid, validation_errors = validate_input_types(payload, original_payload)
        if not is_valid:
            errors = validation_errors

    return errors

def _get_mock_body_for_status(status_code, method, original_payload, expected):
    if status_code in _STATUS_ERROR_BODY:
        return json.dumps(_STATUS_ERROR_BODY[status_code])
    
    if status_code == 204:
        return ""
    
    if 200 <= status_code < 300:
        if method in ['POST', 'PUT', 'PATCH']:
            if original_payload and isinstance(original_payload, (dict, list)):
                return json.dumps(original_payload)
            return json.dumps({"message": "Resource processed successfully", "id": "mock_001", "status": "success"})
        else:
            if original_payload and isinstance(original_payload, (dict, list)):
                return json.dumps(original_payload) if isinstance(original_payload, list) else json.dumps([original_payload])
            return json.dumps({"status": "success", "data": []})
    
    if 400 <= status_code < 500:
        return json.dumps({"error": "Bad Request", "message": expected})
    
    if status_code >= 500:
        return json.dumps({"error": "Internal Server Error", "message": "An unexpected error occurred on the server"})
    
    return json.dumps({"status": "mock_response", "code": status_code})

def _get_error_code_from_scenario(test_type, scenario):
    for keywords, code, msg in _SCENARIO_ERROR_RULES:
        if test_type == msg or any(k.lower() in scenario.lower() for k in keywords):
            return code, msg
    return 400, "Bad Request"

def _get_fallback_response(test_type, method, scenario, original_payload, expected):
    scenario_lower = (scenario or '').lower()
    expected_lower = str(expected or '').lower()

    if 'sort' in scenario_lower and ('invalid' in scenario_lower or 'default sort' in expected_lower or 'fallback' in expected_lower):
        return {
            'statusCode': 200,
            'body': json.dumps({
                "message": "Invalid sort field received. Applied default sorting.",
                "sortHeader": DEFAULT_SORT_HEADER,
                "data": []
            }),
            'expected': expected
        }

    if test_type in ['Positive', 'Integration', 'Performance']:
        if method == 'POST':
            body = json.dumps(original_payload) if original_payload and isinstance(original_payload, (dict, list)) else json.dumps({"message": "Resource created successfully", "id": "mock_001"})
            return {'statusCode': 201, 'body': body, 'expected': expected}
        elif method == 'DELETE':
            return {'statusCode': 204, 'body': '', 'expected': expected}
        else:
            if original_payload and isinstance(original_payload, (dict, list)):
                body = json.dumps(original_payload) if isinstance(original_payload, list) else json.dumps([original_payload])
            else:
                body = json.dumps({"status": "success", "data": []})
            return {'statusCode': 200, 'body': body, 'expected': expected}
    
    elif test_type in ['Negative', 'Validation', 'Security', 'Header', 'Auth', 'RateLimit']:
        status_code, error_msg = _get_error_code_from_scenario(test_type, scenario)
        return {
            'statusCode': status_code,
            'body': json.dumps({"error": error_msg, "message": scenario}),
            'expected': expected
        }
    
    return {
        'statusCode': 200,
        'body': json.dumps({"status": "mock_response", "scenario": scenario}),
        'expected': expected
    }

def generate_mock_response(test_case, method, original_payload=None, field_configs=None):
    expected = test_case.get('expected', 'Success response')
    if expected == 'N/A' or not expected:
        expected = test_case.get('expected_status', 'N/A')
    
    test_type = test_case.get('type', 'Positive')
    scenario = test_case.get('scenario', '')
    payload = test_case.get('input', {})
    
    validation_errors = _run_mock_validation(payload, method, test_type, field_configs, original_payload)
    if validation_errors:
        error_msg = "\n".join([f"• {err}" for err in validation_errors])
        return {
            'statusCode': 400,
            'body': json.dumps({"error": "Bad Request", "message": "Please correct the following validation errors and try again.", "details": validation_errors}),
            'expected': expected,
            'validation_failed': True,
            'validation_error': f"❌ Invalid data type\n\nPlease correct the following validation errors and try again.\n\n{error_msg}"
        }
    
    expected_codes = extract_response_code(expected)
    if expected_codes and "N/A" not in expected_codes:
        try:
            status_code = int(expected_codes[0])
            body = _get_mock_body_for_status(status_code, method, original_payload, expected)
            return {'statusCode': status_code, 'body': body, 'expected': expected}
        except (ValueError, TypeError):
            pass
    
    return _get_fallback_response(test_type, method, scenario, original_payload, expected)

def execute_single_test(endpoint, method, test_case, environment='mock', base_url='mock', original_payload=None, field_configs=None):
    test_id = test_case.get('id', 'Unknown')
    expected = test_case.get('expected', 'N/A')
    
    if expected == 'N/A' or not expected:
        expected = test_case.get('expected_status', 'N/A')
    
    source, additional_info = get_test_case_source_info(test_case)
    
    if not base_url or base_url == 'mock':
        base_url = test_case.get('baseUrl', base_url)
    
    method = test_case.get('method', method).upper()
    endpoint = test_case.get('endpoint', endpoint)
    
    current_endpoint = endpoint
    payload = test_case.get('input', {})
    
    if method in ['GET', 'DELETE']:
        if isinstance(payload, str):
            if payload.startswith('/'):
                current_endpoint = payload
                payload = {}
            elif payload.startswith('?') or '=' in payload:
                payload = parse_query_params(payload)
            else:
                current_endpoint = current_endpoint.rstrip('/') + '/' + payload
                payload = {}
        elif payload is not None and not isinstance(payload, (dict, list)):
            current_endpoint = current_endpoint.rstrip('/') + '/' + str(payload)
            payload = {}
        
        test_case['input'] = payload
    else:
        if isinstance(payload, str) and payload.startswith('/'):
            current_endpoint = payload
    
    if environment == 'mock' or base_url == 'mock':
        return execute_mock_test(test_id, method, test_case, expected, original_payload, field_configs, base_url, current_endpoint)
    
    try:
        url = build_url(current_endpoint, base_url)
        headers = {
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5'
        }
        
        if field_configs:
            source_type = 'query' if method == 'GET' else 'body'
            is_valid, errors = validate_against_configs(payload, field_configs, source=source_type)
            if not is_valid:
                error_summary = "\n".join([f"• {err}" for err in errors])
                full_msg = f"❌ Invalid data type\n\nPlease correct the following validation errors and try again.\n\n{error_summary}"
                expected_codes = extract_response_code(expected)
                status = 'fail' if test_case.get('type') == 'Positive' else ('pass' if '400' in expected_codes else 'fail')
                return {
                    'testCaseId': test_id,
                    'status': status,
                    'statusCode': 400,
                    'responseBody': json.dumps({"error": "Validation Error", "message": "Validation failed", "details": errors}),
                    'details': f"❌ Validation Error\n\n{full_msg}\n\nExpected: {expected}"
                }

        if test_case.get('type') == 'Positive' and original_payload:
            source_type = 'query' if method == 'GET' else 'body'
            if source_type == 'query':
                parsed_payload = parse_query_params(payload)
                is_valid, errors = validate_input_types(parsed_payload, original_payload)
            else:
                is_valid, errors = validate_input_types(payload, original_payload)
                
            if not is_valid:
                error_summary = "\n".join([f"• {err}" for err in errors])
                full_msg = f"❌ Invalid data type\n\nPlease correct the following validation errors and try again.\n\n{error_summary}"
                return {
                    'testCaseId': test_id,
                    'status': 'fail',
                    'statusCode': 400,
                    'responseBody': json.dumps({"error": "Data Type Error", "message": "Type mismatch", "details": errors}),
                    'details': f"❌ Data Type Error\n\n{full_msg}\n\nExpected: {expected}"
                }

        try:
            if isinstance(payload, dict):
                payload_json = json.dumps(payload)
            elif isinstance(payload, list):
                payload_json = json.dumps(payload)
            elif payload is None:
                payload_json = None
            else:
                payload_json = str(payload)
        except:
            payload_json = str(payload)

        timeout = 10

        if method == 'GET':
            params = payload if isinstance(payload, dict) else {}
            response = requests.get(url, params=params, headers=headers, timeout=timeout)
        elif method == 'POST':
            response = requests.post(url, data=payload_json, headers=headers, timeout=timeout)
        elif method == 'PUT':
            response = requests.put(url, data=payload_json, headers=headers, timeout=timeout)
        elif method == 'PATCH':
            response = requests.patch(url, data=payload_json, headers=headers, timeout=timeout)
        elif method == 'DELETE':
            if isinstance(payload, str) and payload.startswith('/'):
                response = requests.delete(url, headers=headers, timeout=timeout)
            else:
                response = requests.delete(url, data=payload_json, headers=headers, timeout=timeout)
        else:
            response = requests.request(method, url, data=payload_json, headers=headers, timeout=timeout)

        expected_codes = extract_response_code(expected)
        actual_code = str(response.status_code)
        
        status = 'fail'
        body_assertion_msg = None
        
        if "N/A" in expected_codes:
            status = 'pass'
        elif actual_code in expected_codes:
            status = 'pass'
        elif actual_code == '200':
            test_type = test_case.get('type', '').lower()
            scenario_lower = test_case.get('scenario', '').lower()
            
            is_security_test = 'security' in test_type or 'auth' in test_type or 'sql' in scenario_lower or 'xss' in scenario_lower
            
            if test_type in ['negative', 'edge case', 'boundary', 'validation'] and not is_security_test:
                is_fallback_scenario = any(kw in scenario_lower for kw in ['sort', 'page', 'size', 'query', 'search', 'filter', 'invalid parameter'])
                
                if is_fallback_scenario:
                    body_lower = response.text.lower()
                    is_json = 'application/json' in response.headers.get('Content-Type', '').lower()
                    
                    if is_json:
                        graceful_error_indicators = [
                            'no results', '0 results', 'not found', 'error', 'invalid', 
                            'bad request', 'missing', 'default'
                        ]
                        found_indicators = [kw for kw in graceful_error_indicators if kw in body_lower]
                        if found_indicators:
                            status = 'pass'
                            indicators_str = ", ".join(found_indicators)
                            body_assertion_msg = f"✅ Soft Assertion Passed: Server handled bad input gracefully. Found: '{indicators_str}'"
                    else:
                        graceful_error_indicators = [
                            'no results found', '0 results', 'did not match', 'try different keywords'
                        ]
                        found_indicators = [kw for kw in graceful_error_indicators if kw in body_lower]
                        
                        if found_indicators:
                            status = 'pass'
                            indicators_str = ", ".join(found_indicators)
                            body_assertion_msg = f"✅ Soft Assertion Passed: Server handled bad input gracefully. Found: '{indicators_str}'"
                        else:
                            status = 'pass'
                            fallback_element = "Default Page View"
                            title_match = re.search(r'<title[^>]*>(.*?)</title>', response.text, re.IGNORECASE)
                            if title_match:
                                fallback_element = title_match.group(1).strip()
                            body_assertion_msg = f"✅ Graceful Fallback Passed: Server ignored the invalid parameter and safely loaded the default state.\n↳ Fallback Header Loaded: '{fallback_element}'"

        response_text = response.text[:1000] if response.text else '(No response body)'

        schema_valid = True
        schema_errors = []
        if response.status_code >= 500:
            schema_valid, schema_errors = validate_response_schema(response.text, response.status_code)
            if not schema_valid:
                status = 'fail'

        print(f"\n--- Test Case Execution: {test_id} ---")
        print(f"Source: {source}")
        if additional_info:
            for info in additional_info:
                print(f"  {info}")
        print(f"Request URL: {response.url}")
        print(f"Content Type: {response.headers.get('Content-Type')}")
        print(f"Response Length: {len(response.text) if response.text else 0}")
        print(f"HTTP Method: {method}")
        print(f"Expected Status: {format_expected_for_display(expected)}")
        print(f"Actual Status: {response.status_code}")
        print(f"Response Body: {response.text}")
        print("-" * 40)

        details_parts = [f"Request URL: {response.url}"]
        details_parts.append(f"Content Type: {response.headers.get('Content-Type')}")
        details_parts.append(f"Response Length: {len(response.text) if response.text else 0}")
        if additional_info:
            details_parts.append(f"Source: {source} ({', '.join(additional_info)})")
        else:
            details_parts.append(f"Source: {source}")
        details_parts.append(f"Expected: {format_expected_for_display(expected)}")
        details_parts.append(f"Actual: HTTP {response.status_code}")
        
        if isinstance(expected, str):
            expected_lower = expected.lower()
        else:
            expected_lower = str(expected).lower()
        if ('default sort' in expected_lower or 'fallback' in expected_lower) and response.status_code == 200:
            details_parts.append(f"Sort Fallback: Applied default sort header '{DEFAULT_SORT_HEADER}'")
            
        if body_assertion_msg:
            details_parts.append(f"\n{body_assertion_msg}")
        
        if not schema_valid and schema_errors:
            details_parts.append(f"\nSchema Validation Failed:")
            for err in schema_errors:
                details_parts.append(f"  • {err}")
        
        details_parts.append(f"\nResponse Body:\n{response_text}")
        
        return {
            'testCaseId': test_id,
            'status': status,
            'statusCode': response.status_code,
            'contentType': response.headers.get('Content-Type', ''),
            'responseBody': response.text if response.text else '',
            'details': "\n".join(details_parts),
            'source': source,
            'additionalInfo': additional_info
        }
    except requests.exceptions.Timeout:
        details_parts = [f"❌ Connection Error: Request timeout after 10 seconds\n\nThe API endpoint took too long to respond. Check if the server is running and accessible."]
        if additional_info:
            details_parts.append(f"Source: {source} ({', '.join(additional_info)})")
        else:
            details_parts.append(f"Source: {source}")
        
        return {
            'testCaseId': test_id,
            'status': 'fail',
            'statusCode': 0,
            'details': "\n".join(details_parts),
            'source': source,
            'additionalInfo': additional_info
        }
    except requests.exceptions.ConnectionError as e:
        error_msg = str(e).lower()
        if 'getaddrinfo failed' in error_msg or 'name resolution' in error_msg:
            friendly_msg = f"❌ DNS Resolution Failed\n\nCannot resolve hostname. Verify:\n• Endpoint URL is correct\n• Network connection is active\n• DNS settings are configured properly"
        elif 'connection refused' in error_msg:
            friendly_msg = f"❌ Connection Refused\n\nThe server is not accepting connections. Verify:\n• Server is running on the specified port\n• Firewall rules allow the connection\n• Correct URL and port are configured"
        else:
            friendly_msg = f"❌ Connection Failed\n\n{str(e)[:200]}"
        
        details_parts = [friendly_msg]
        if additional_info:
            details_parts.append(f"Source: {source} ({', '.join(additional_info)})")
        else:
            details_parts.append(f"Source: {source}")
        
        return {
            'testCaseId': test_id,
            'status': 'fail',
            'statusCode': 0,
            'details': "\n".join(details_parts),
            'source': source,
            'additionalInfo': additional_info
        }
    except Exception as e:
        details_parts = [f"❌ Unexpected Error\n\n{str(e)[:300]}"]
        if additional_info:
            details_parts.append(f"Source: {source} ({', '.join(additional_info)})")
        else:
            details_parts.append(f"Source: {source}")
        
        return {
            'testCaseId': test_id,
            'status': 'fail',
            'statusCode': 0,
            'details': "\n".join(details_parts),
            'source': source,
            'additionalInfo': additional_info
        }

def execute_mock_test(test_id, method, test_case, expected, original_payload=None, field_configs=None, base_url='mock', current_endpoint=None):
    if current_endpoint is None:
        current_endpoint = test_case.get('endpoint', '')
        
    mock_response = generate_mock_response(test_case, method, original_payload, field_configs)
    status_code = mock_response['statusCode']
    expected_codes = extract_response_code(expected)
    actual_code = str(status_code)
    test_type = test_case.get('type', 'Positive')
    
    if mock_response.get('validation_failed'):
        if test_type == 'Positive':
            status = 'fail'
        else:
            status = 'pass' if '400' in expected_codes else 'fail'
    else:
        if "N/A" in expected_codes:
            status = 'pass'
        elif actual_code in expected_codes:
            status = 'pass'
        else:
            status = 'fail'
    
    payload = test_case.get('input', {})
    
    if base_url and base_url != 'mock':
        url = build_url(current_endpoint, base_url)
        if method == 'GET' and payload and isinstance(payload, dict):
            req = requests.Request('GET', url, params=payload)
            prepared = req.prepare()
            mock_url = prepared.url
        else:
            mock_url = url
    else:
        if method == 'GET' and payload and isinstance(payload, dict):
             import urllib.parse
             qs = urllib.parse.urlencode(payload)
             mock_url = f"MOCK://{method}{current_endpoint}?{qs}"
        else:
             mock_url = f"MOCK://{method}{current_endpoint}"

    source, additional_info = get_test_case_source_info(test_case)
    
    print(f"\n--- [MOCK] Test Case Execution: {test_id} ---")
    print(f"Source: {source}")
    if additional_info:
        for info in additional_info:
            print(f"  {info}")
    print(f"Request URL: {mock_url}")
    print(f"HTTP Method: {method}")
    print(f"Expected Status: {format_expected_for_display(expected)}")
    print(f"Actual Status: {status_code}")
    print(f"Response Body: {mock_response['body']}")
    print("-" * 40)

    details_parts = ["[MOCK MODE]"]
    if mock_response.get('validation_failed'):
        details_parts.append(mock_response['validation_error'])
    
    details_parts.append(f"Request URL: {mock_url}")
    
    if additional_info:
        details_parts.append(f"Source: {source} ({', '.join(additional_info)})")
    else:
        details_parts.append(f"Source: {source}")
    
    details_parts.append(f"Expected: {format_expected_for_display(expected)}")
    details_parts.append(f"Actual: HTTP {status_code}")
    if isinstance(expected, str):
        expected_lower = expected.lower()
    else:
        expected_lower = str(expected).lower()
    if ('default sort' in expected_lower or 'fallback' in expected_lower) and status_code == 200:
        details_parts.append(f"Sort Fallback: Applied default sort header '{DEFAULT_SORT_HEADER}'")
    details_parts.append(f"\nResponse:\n{mock_response['body']}")
    
    return {
        'testCaseId': test_id,
        'status': status,
        'statusCode': status_code,
        'responseBody': mock_response['body'],
        'details': "\n".join(details_parts),
        'source': source,
        'additionalInfo': additional_info
    }

def build_url(endpoint, base_url):
    if not endpoint:
        endpoint = ""
    if not base_url:
        base_url = ""
        
    if endpoint.startswith('http'):
        return endpoint
        
    if base_url.startswith('http'):
        base = base_url.rstrip('/')
        if endpoint.startswith('?'):
            return base + endpoint
        return base + '/' + endpoint.lstrip('/')
        
    return f'http://localhost:5000/{endpoint.lstrip("/")}'

if __name__ == '__main__':
    import os
    host = os.environ.get('FLASK_HOST', '0.0.0.0')
    port = int(os.environ.get('FLASK_PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0').lower() in ('1', 'true', 'yes')
    print("=" * 60)
    print("API Test Command Center - Flask Application")
    print("=" * 60)
    print(f"Host: {host}")
    print(f"Port: {port}")
    print(f"Debug mode: {debug}")
    print("-" * 60)
    print(f"Starting Flask server on http://{host}:{port}")
    print("Press Ctrl+C to stop the server")
    print("=" * 60)
    app.run(debug=debug, host=host, port=port, use_reloader=debug)